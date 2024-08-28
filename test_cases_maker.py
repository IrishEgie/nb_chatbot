import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation

# Load the text file
file_path = 'test_cases.txt'
with open(file_path, 'r') as file:
    content = file.read().splitlines()

# Initialize workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Test Cases"

# Set headers
headers = ["Select", "Question", "Correct Prognosis", "Bot Answer"]
for col_num, header in enumerate(headers, 1):
    worksheet.cell(row=1, column=col_num, value=header)

# Create checkboxes and populate questions and prognosis
row = 2
for line in content:
    if "**Question**" in line:
        question = line.split(": ", 1)[1]
        worksheet.cell(row=row, column=2, value=question)
    elif "**Prognosis**" in line:
        prognosis = line.split(": ", 1)[1]
        worksheet.cell(row=row, column=3, value=prognosis)
        row += 1

# Add checkboxes to the first column
for row in range(2, worksheet.max_row + 1):
    worksheet.cell(row=row, column=1, value="☐")  # Add a checkbox symbol

# Adjust column widths
for col in range(1, worksheet.max_column + 1):
    worksheet.column_dimensions[get_column_letter(col)].width = 20

# Save the workbook
output_file = 'Chatbot_Test_Cases.xlsx'
workbook.save(output_file)
print(f"Excel file created and saved as {output_file}")
